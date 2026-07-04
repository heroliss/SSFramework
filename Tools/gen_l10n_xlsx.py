# 一次性脚本：生成 demo 的 l10n.xlsx（Luban 多语言表数据源）。
# 注意：从项目根运行，输出路径相对项目根。
# 布局遵循 Luban 约定：A 列为标记列，##var 行写字段名，## 行为注释行，数据行 A 列留空。
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "l10n"

rows = [
    ["##var", "key", "zhCn", "en"],
    ["##", "文案key(主键)", "简体中文", "English(留空=走fallback)"],
    [None, "menu/start", "开始游戏", "Start Game"],
    [None, "lobby/welcome", "欢迎回来，{0}！", "Welcome back, {0}!"],
    [None, "demo/clicks", "已点击 {0} 次", "Clicked {0} time(s)"],
    [None, "demo/only-zh", "这条文案只有中文（英文下走 fallback 仍显示我）", None],
    [None, "l10n/from-table", "这行文本来自 Luban 表 TbL10N（Excel 一行一 key、一列一语言）", "This line comes from Luban table TbL10N (one key per row, one language per column)"],
    [None, "l10n/banner-caption", "下图按当前语言加载：l10n-banner_{0}", "Image below is loaded per locale: l10n-banner_{0}"],
    [None, "l10n/voice-caption", "播放当前语言的提示音：l10n-voice_{0}", "Play the voice cue for current locale: l10n-voice_{0}"],
]
for r in rows:
    ws.append(r)

header_fill = PatternFill("solid", fgColor="D9E2F3")
for row_idx in (1, 2):
    for cell in ws[row_idx]:
        cell.font = Font(bold=(row_idx == 1))
        cell.fill = header_fill
widths = {"A": 8, "B": 22, "C": 52, "D": 60}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

_out = "Assets/Game/Framework/Demo/Configs~/Datas/l10n.xlsx"
wb.save(_out)
print("written:", _out)
