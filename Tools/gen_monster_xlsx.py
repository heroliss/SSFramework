# 一次性脚本：生成 Configs/Datas/monster.xlsx（Luban Excel 数据源示例）。
# 布局遵循 Luban 约定：A 列为标记列，##var 行写字段名，## 行为注释行，数据行 A 列留空。
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "monster"

rows = [
    ["##var", "id", "name", "hp", "attack", "dropItemId"],
    ["##", "怪物ID(主键)", "名称", "生命值", "攻击力", "掉落物品id(对应TbItem.id)"],
    [None, 101, "史莱姆", 30, 5, 2001],
    [None, 102, "哥布林", 80, 12, 1001],
    [None, 103, "骷髅弓手", 120, 18, 1002],
    [None, 201, "石像守卫", 600, 45, 2002],
    [None, 301, "幼龙", 2400, 130, 3001],
    [None, 999, "堕落神祇", 99999, 999, 9001],
]
for r in rows:
    ws.append(r)

# 表头加粗 + 底色，列宽放宽——给真人用 Excel/WPS 打开编辑时的可读性
header_fill = PatternFill("solid", fgColor="D9E2F3")
for row_idx in (1, 2):
    for cell in ws[row_idx]:
        cell.font = Font(bold=(row_idx == 1))
        cell.fill = header_fill
widths = {"A": 8, "B": 8, "C": 14, "D": 8, "E": 8, "F": 26}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save("Configs/Datas/monster.xlsx")
print("written: Configs/Datas/monster.xlsx")
